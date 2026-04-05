"""
Goofish / Xianyu - Scraper de Vendedor
========================================
Como usar:
  python goofish_scraper.py

O script vai te pedir:
  1. ID ou link do vendedor
  2. Palavra-chave para filtrar (ex: 苹果13 美版)

Resultado: salva goofish_[parte latina da palavra-chave].xlsx na mesma pasta
"""

import asyncio
import re
import sys
from datetime import datetime
from pathlib import Path

# Verifica dependências antes de importar
try:
    from playwright.async_api import async_playwright
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError as e:
    print(f"\n❌ Dependência faltando: {e}")
    print("Rode primeiro: python -m pip install playwright openpyxl")
    sys.exit(1)


# ─── Extrai o ID do vendedor a partir de um link ou ID puro ───────────────────

def extrair_seller_id(entrada: str) -> str:
    entrada = entrada.strip()
    # Se for link, tenta extrair o userId
    padrao = re.search(r'userId[=/](\d+)', entrada)
    if padrao:
        return padrao.group(1)
    # Se for só números, assume que é o ID direto
    if entrada.isdigit():
        return entrada
    # Tenta pegar qualquer sequência de dígitos longa no link
    numeros = re.findall(r'\d{6,}', entrada)
    if numeros:
        return numeros[0]
    return entrada  # devolve como está e tenta assim mesmo


# ─── Gera nome do arquivo a partir da palavra-chave ──────────────────────────

def nome_arquivo(palavra_chave: str) -> str:
    # Pega só letras latinas (a-z A-Z) e números (0-9), ignora caracteres chineses e espaços
    parte_latina = re.sub(r'[^a-zA-Z0-9]+', '_', palavra_chave).strip('_')
    if not parte_latina:
        parte_latina = "resultado"
    return f"goofish_{parte_latina}.xlsx"


# ─── Scraping principal ───────────────────────────────────────────────────────

async def scraper(seller_id: str, palavra_chave: str) -> list[dict]:
    url_perfil = f"https://www.goofish.com/personal?userId={seller_id}"
    print(f"\n🌐 Abrindo perfil: {url_perfil}")
    print("⏳ Aguarde, isso pode levar alguns segundos...\n")

    resultados = []

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=False,  # False = abre janela visível (mais difícil de bloquear)
            args=["--lang=zh-CN"]
        )

        context = await browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            locale="zh-CN",
            viewport={"width": 1280, "height": 800},
        )

        page = await context.new_page()

        # Vai para o perfil do vendedor
        try:
            await page.goto(url_perfil, timeout=30000, wait_until="domcontentloaded")
        except Exception as e:
            print(f"❌ Erro ao abrir o perfil: {e}")
            await browser.close()
            return []

        # Aguarda a página carregar os itens
        await page.wait_for_timeout(4000)

        # Fecha o pop-up de login automaticamente
        try:
            seletores_fechar = [
                'button[aria-label="Close"]',
                'button[aria-label="close"]',
                '.close-btn',
                '[class*="closeBtn"]',
                '[class*="CloseBtn"]',
                '[class*="close-icon"]',
            ]
            fechou = False
            for seletor in seletores_fechar:
                botao_x = await page.query_selector(seletor)
                if botao_x:
                    await botao_x.click()
                    fechou = True
                    break

            if not fechou:
                # Tenta pressionar ESC para fechar o modal
                await page.keyboard.press('Escape')

            await page.wait_for_timeout(1000)
            print("✅ Pop-up de login fechado.")
        except Exception:
            pass  # Se não tiver pop-up, continua normalmente

        # Tenta clicar no botão "在售" pelo TEXTO, independente da posição na página
        try:
            # Busca por divs com classe "tabItem" que contenham o texto 在售
            handles = await page.query_selector_all('[class*="tabItem"]')
            botao_em_venda = None
            for el in handles:
                txt = await el.inner_text()
                if '在售' in txt:
                    botao_em_venda = el
                    break

            # Fallback: varre span/div/button procurando texto que contenha 在售
            if not botao_em_venda:
                handles = await page.query_selector_all('span, div, button')
                for el in handles:
                    txt = await el.inner_text()
                    if txt.strip().startswith('在售') or txt.strip() == '在售':
                        botao_em_venda = el
                        break

            if botao_em_venda:
                await botao_em_venda.click()
                print("✅ Filtro '在售' (em venda) ativado — ignorando anúncios já vendidos.")
                await page.wait_for_timeout(2000)
            else:
                print("⚠️  Botão '在售' não encontrado — coletando todos os anúncios.")
        except Exception:
            print("⚠️  Não foi possível ativar o filtro '在售' — coletando todos os anúncios.")

        pagina_atual = 1
        max_paginas = 50  # segurança para não rodar infinito

        while pagina_atual <= max_paginas:
            print(f"📄 Lendo página {pagina_atual}...")

            # Scroll para carregar mais itens (página infinita)
            await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            await page.wait_for_timeout(2000)
            await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            await page.wait_for_timeout(1500)

            # Tenta coletar os cards de produto
            # O Goofish usa classes dinâmicas, então buscamos pelos links de produto
            cards = await page.query_selector_all('a[href*="/item/"]')

            if not cards:
                # Tenta seletor alternativo
                cards = await page.query_selector_all('a[href*="item"]')

            novos = 0
            ids_vistos = {r["_href"] for r in resultados}

            for card in cards:
                try:
                    href = await card.get_attribute("href")
                    if not href:
                        continue

                    # Monta URL completa
                    if href.startswith("http"):
                        link = href
                    else:
                        link = "https://www.goofish.com" + href

                    if link in ids_vistos:
                        continue

                    # Tenta pegar o texto do card (título + preço)
                    texto_completo = await card.inner_text()
                    linhas = [l.strip() for l in texto_completo.split("\n") if l.strip()]

                    titulo = ""
                    preco = ""

                    for linha in linhas:
                        # Linha com preço: contém ¥ ou só dígitos com ponto
                        if "¥" in linha or re.match(r'^\d+\.?\d*$', linha):
                            preco = linha.replace("¥", "¥").strip()
                        elif len(linha) > 3 and not preco:
                            titulo = linha

                    # Filtro por tokens: separa a palavra-chave em partes
                    # Ex: "苹果16无锁" → ["苹果", "16", "无锁"]
                    # Tokens numéricos (ex: "16") só batem com o TÍTULO, não com o preço
                    # e precisam aparecer como número isolado (não dentro de "1659" ou "160")
                    if palavra_chave:
                        titulo_busca = titulo.lower()
                        tudo_busca   = (titulo + texto_completo).lower()
                        tokens = re.findall(r"[a-zA-Z0-9]+|[一-鿿]+", palavra_chave)

                        token_ok = True
                        for t in tokens:
                            t_lower = t.lower()
                            if t.isdigit():
                                # Número puro: só busca no título e como palavra isolada
                                # ex: "16" não pode estar dentro de "1659" ou "160"
                                if not re.search(r'(?<!\d)' + re.escape(t_lower) + r'(?!\d)', titulo_busca):
                                    token_ok = False
                                    break
                            else:
                                # Texto normal: busca em tudo (título + texto completo)
                                if t_lower not in tudo_busca:
                                    token_ok = False
                                    break

                        if not token_ok:
                            continue

                    if titulo or preco:
                        resultados.append({
                            "Título": titulo or texto_completo[:60],
                            "Preço": preco or "—",
                            "Link": link,
                            "_href": link,
                        })
                        ids_vistos.add(link)
                        novos += 1

                except Exception:
                    continue

            print(f"   ✅ {novos} novos anúncios encontrados (total: {len(resultados)})")

            # Tenta clicar em "próxima página" ou detecta fim do scroll
            proximo = await page.query_selector('[class*="next"], [aria-label*="next"], button:has-text("下一页")')
            if proximo:
                try:
                    await proximo.click()
                    await page.wait_for_timeout(3000)
                    pagina_atual += 1
                except Exception:
                    break
            else:
                # Verifica se ainda tem conteúdo novo depois do scroll
                altura_antes = await page.evaluate("document.body.scrollHeight")
                await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                await page.wait_for_timeout(2500)
                altura_depois = await page.evaluate("document.body.scrollHeight")

                if altura_depois <= altura_antes:
                    print("\n📌 Fim da página — sem mais itens para carregar.")
                    break

                pagina_atual += 1

        await browser.close()

    # Remove campo auxiliar
    for r in resultados:
        r.pop("_href", None)

    return resultados


# ─── Salvar Excel ─────────────────────────────────────────────────────────────

def extrair_preco_numerico(item: dict) -> float:
    # Remove tudo que não for dígito ou ponto e converte para float
    preco_str = re.sub(r'[^\d.]', '', item.get("Preço", ""))
    try:
        return float(preco_str)
    except ValueError:
        return float('inf')  # Itens sem preço vão para o final


def salvar_excel(dados: list[dict], caminho: Path):
    # Ordena por preço crescente antes de salvar
    dados = sorted(dados, key=extrair_preco_numerico)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Anúncios"

    cabecalhos = ["Título", "Preço", "Link"]
    cor_cabecalho = PatternFill("solid", fgColor="1A1A2E")
    fonte_cabecalho = Font(color="FFFFFF", bold=True, size=11)

    for col, cab in enumerate(cabecalhos, 1):
        cell = ws.cell(row=1, column=col, value=cab)
        cell.fill = cor_cabecalho
        cell.font = fonte_cabecalho
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Largura das colunas
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 60

    for i, item in enumerate(dados, 2):
        ws.cell(row=i, column=1, value=item.get("Título", ""))
        ws.cell(row=i, column=2, value=item.get("Preço", ""))

        link = item.get("Link", "")
        cell_link = ws.cell(row=i, column=3, value=link)
        cell_link.hyperlink = link
        cell_link.font = Font(color="0563C1", underline="single")

        if i % 2 == 0:
            for col in range(1, 4):
                ws.cell(row=i, column=col).fill = PatternFill("solid", fgColor="F5F5F5")

    wb.save(caminho)


# ─── Entrada do usuário ───────────────────────────────────────────────────────

def pedir_entrada():
    print("=" * 55)
    print("  🐟  GOOFISH SCRAPER — Buscador de Vendedor")
    print("=" * 55)
    print()

    entrada = input("📌 Cole o ID ou o link do perfil do vendedor:\n> ").strip()
    seller_id = extrair_seller_id(entrada)
    print(f"   → ID detectado: {seller_id}")

    print()
    palavra = input("🔍 Palavra-chave para filtrar (deixe vazio para pegar tudo):\n> ").strip()

    return seller_id, palavra


# ─── Main ─────────────────────────────────────────────────────────────────────

async def main():
    seller_id, palavra_chave = pedir_entrada()

    if not seller_id:
        print("❌ ID inválido. Tente novamente.")
        sys.exit(1)

    dados = await scraper(seller_id, palavra_chave)

    if not dados:
        print("\n⚠️  Nenhum anúncio encontrado.")
        print("Possíveis causas:")
        print("  • O ID do vendedor está errado")
        print("  • O site bloqueou a requisição (tente novamente)")
        print("  • A palavra-chave não bateu com nenhum título")
        input("\nPressione Enter para fechar...")
        sys.exit(0)

    pasta = Path.cwd()
    arquivo_xlsx = pasta / nome_arquivo(palavra_chave if palavra_chave else seller_id)

    salvar_excel(dados, arquivo_xlsx)

    print()
    print("=" * 55)
    print("  ✅  CONCLUÍDO!")
    print("=" * 55)
    print(f"  📦 Anúncios encontrados : {len(dados)}")
    print(f"  📊 Arquivo              : {arquivo_xlsx.name}")
    print(f"  📁 Pasta                : {pasta}")
    print("=" * 55)

    # Mostra prévia no terminal
    print()
    print(f"{'TÍTULO':<35} {'PREÇO':>10}")
    print("─" * 55)
    for item in dados[:10]:
        titulo = item['Título'][:33] + ".." if len(item['Título']) > 35 else item['Título']
        print(f"{titulo:<35} {item['Preço']:>10}")
    if len(dados) > 10:
        print(f"  ... e mais {len(dados) - 10} anúncios no arquivo.")
    print("─" * 55)
    print()
    input("Pressione Enter para fechar...")


if __name__ == "__main__":
    asyncio.run(main())
